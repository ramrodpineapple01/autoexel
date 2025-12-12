"""
Excel file handler for WCCSA Community Directory Management Tool
Handles all Excel operations using openpyxl
"""
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelHandler:
    def __init__(self, file_path='data/community_data.xlsx'):
        self.file_path = file_path
        self.ensure_data_directory()
        self.init_workbook()
    
    def ensure_data_directory(self):
        """Create data directory if it doesn't exist"""
        os.makedirs(os.path.dirname(self.file_path) if os.path.dirname(self.file_path) else '.', exist_ok=True)
    
    def init_workbook(self):
        """Initialize or load the Excel workbook"""
        if os.path.exists(self.file_path):
            self.wb = load_workbook(self.file_path)
        else:
            self.wb = Workbook()
            # Remove default sheet
            if 'Sheet' in self.wb.sheetnames:
                self.wb.remove(self.wb['Sheet'])
            self.create_sheets()
            self.wb.save(self.file_path)
    
    def create_sheets(self):
        """Create all required sheets with headers"""
        # Sheet 1: Directory
        if 'Directory' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Directory')
            headers = ['ID', 'Owner', 'Phone', 'Address', 'City', 'Zip', 'Email']
            ws.append(headers)
            self.format_headers(ws)
        
        # Sheet 2: Board_of_Directors
        if 'Board_of_Directors' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Board_of_Directors')
            headers = ['Year', 'Position', 'Name', 'Additional_Duties', 'Contact_Info']
            ws.append(headers)
            self.format_headers(ws)
        
        # Sheet 3: Committees
        if 'Committees' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Committees')
            headers = ['Committee_Name', 'Member_Name', 'Role', 'Contact_Info', 'Meeting_Notes']
            ws.append(headers)
            self.format_headers(ws)
        
        # Sheet 4: Lot_Owners
        if 'Lot_Owners' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Lot_Owners')
            headers = ['Surname', 'FirstName', 'Lot_Numbers']
            ws.append(headers)
            self.format_headers(ws)
        
        # Sheet 5: Lot_Map_Regions
        if 'Lot_Map_Regions' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Lot_Map_Regions')
            headers = ['Lot_Number', 'Owner_Name', 'Region_Type', 'Coordinates', 'Label_X', 'Label_Y']
            ws.append(headers)
            self.format_headers(ws)
    
    def format_headers(self, ws):
        """Format header row"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_sheet_data(self, sheet_name):
        """Get all data from a sheet (excluding headers)"""
        if sheet_name not in self.wb.sheetnames:
            return []
        
        ws = self.wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        data = []
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_data = {}
            for idx, cell in enumerate(row):
                row_data[headers[idx]] = cell.value if cell.value is not None else ''
            if any(row_data.values()):  # Only add non-empty rows
                data.append(row_data)
        
        return data
    
    def add_row(self, sheet_name, data):
        """Add a new row to a sheet"""
        if sheet_name not in self.wb.sheetnames:
            return False
        
        ws = self.wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        # Generate ID for Directory sheet
        if sheet_name == 'Directory' and 'ID' in headers:
            max_id = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and isinstance(row[0], (int, str)):
                    try:
                        max_id = max(max_id, int(str(row[0])))
                    except:
                        pass
            data['ID'] = max_id + 1
        
        row_values = [data.get(header, '') for header in headers]
        ws.append(row_values)
        self.wb.save(self.file_path)
        return True
    
    def update_row(self, sheet_name, row_id, data):
        """Update a row in a sheet"""
        if sheet_name not in self.wb.sheetnames:
            return False
        
        ws = self.wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        # Find the row
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if sheet_name == 'Directory' and row[0].value == row_id:
                for header, value in data.items():
                    if header in headers:
                        col_idx = headers.index(header)
                        ws.cell(row=idx, column=col_idx + 1, value=value)
                self.wb.save(self.file_path)
                return True
        
        return False
    
    def delete_row(self, sheet_name, row_id):
        """Delete a row from a sheet"""
        if sheet_name not in self.wb.sheetnames:
            return False
        
        ws = self.wb[sheet_name]
        
        # Find and delete the row
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if sheet_name == 'Directory' and row[0].value == row_id:
                ws.delete_rows(idx)
                self.wb.save(self.file_path)
                return True
        
        return False
    
    def search_directory(self, query):
        """Search directory by any field"""
        data = self.get_sheet_data('Directory')
        query_lower = query.lower()
        
        results = []
        for row in data:
            for value in row.values():
                if query_lower in str(value).lower():
                    results.append(row)
                    break
        
        return results
    
    def get_committees(self):
        """Get all committees with their members"""
        data = self.get_sheet_data('Committees')
        committees = {}
        
        for row in data:
            committee = row.get('Committee_Name', '')
            if committee not in committees:
                committees[committee] = {
                    'members': [],
                    'meeting_notes': ''
                }
            
            member_info = {
                'name': row.get('Member_Name', ''),
                'role': row.get('Role', 'Member'),
                'contact': row.get('Contact_Info', '')
            }
            committees[committee]['members'].append(member_info)
            
            # Get meeting notes (assuming last one wins, or we could aggregate)
            if row.get('Meeting_Notes'):
                committees[committee]['meeting_notes'] = row.get('Meeting_Notes', '')
        
        return committees
    
    def save_committee(self, committee_name, members, meeting_notes):
        """Save committee data (replace existing)"""
        ws = self.wb['Committees']
        
        # Delete existing rows for this committee
        rows_to_delete = []
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == committee_name:
                rows_to_delete.append(idx)
        
        for idx in reversed(rows_to_delete):
            ws.delete_rows(idx)
        
        # Add new rows
        for member in members:
            ws.append([
                committee_name,
                member.get('name', ''),
                member.get('role', 'Member'),
                member.get('contact', ''),
                meeting_notes if member == members[0] else ''  # Only store notes once
            ])
        
        self.wb.save(self.file_path)
    
    def get_lot_map_regions(self):
        """Get all lot map regions"""
        data = self.get_sheet_data('Lot_Map_Regions')
        return data
    
    def save_lot_map_region(self, lot_number, owner_name, region_type, coordinates, label_x, label_y):
        """Save or update a lot map region"""
        ws = self.wb['Lot_Map_Regions']
        
        # Check if lot already exists
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == lot_number:
                # Update existing
                ws.cell(row=idx, column=1, value=lot_number)
                ws.cell(row=idx, column=2, value=owner_name)
                ws.cell(row=idx, column=3, value=region_type)
                ws.cell(row=idx, column=4, value=json.dumps(coordinates))
                ws.cell(row=idx, column=5, value=label_x)
                ws.cell(row=idx, column=6, value=label_y)
                self.wb.save(self.file_path)
                return
        
        # Add new
        ws.append([
            lot_number,
            owner_name,
            region_type,
            json.dumps(coordinates),
            label_x,
            label_y
        ])
        self.wb.save(self.file_path)
    
    def get_lot_map_region(self, lot_number):
        """Get a specific lot map region"""
        data = self.get_lot_map_regions()
        for region in data:
            if region.get('Lot_Number') == lot_number:
                coords = region.get('Coordinates', '[]')
                if isinstance(coords, str):
                    try:
                        coords = json.loads(coords)
                    except:
                        coords = []
                region['Coordinates'] = coords
                return region
        return None

